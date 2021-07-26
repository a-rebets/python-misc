# used dataset - https://www.kaggle.com/saurograndi/airplane-crashes-since-1908?select=Airplane_Crashes_and_Fatalities_Since_1908.csv

import sys
import os
import argparse
import math
import logging
import csv
import openpyxl
from itertools import groupby
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

READ_MODE = 'r'
logging.basicConfig(
    format='%(asctime)s > %(message)s', datefmt='%d/%m/%Y %I:%M:%S %p',
    filename='output.log', level=logging.DEBUG
)
XLS_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))


def run():
    parser = argparse.ArgumentParser(
        description="An app that analyzes a given flight dataset")
    parser.add_argument('dataset', help='a file containing the input dataset')
    parser.add_argument("-o", "--optional",
                        help="generate Excel report", nargs='*')
    args = parser.parse_args()
    try:
        received = _read_csv(args.dataset.strip())
        if not received:
            raise ValueError('No data found')
        parsed_list = _parse_csv()
        report = []
        report.append(_get_stat_info(parsed_list))
        report.append(_get_aggreg_info(parsed_list))
        report.append(_get_summary_info(parsed_list))
        if args.optional is not None:
            if args.optional:
                os.environ['XLS_NAME'] = args.optional[0]
            _generate_xls_report(report)
        else:
            _print_report(report[2])
    except ValueError as e:
        logging.exception(str(e))
    except AssertionError as e:
        logging.exception(e)


def _read_csv(filename):
    data = None
    assert filename.endswith('.csv'), 'The file type should be CSV!'
    os.environ['DATAFILE'] = filename
    try:
        with open(filename) as fobj:
            data = fobj.read()
    except IOError:
        logging.exception('Could not read a file at the given path!')
    return data


def _parse_csv():
    with open(os.environ.get('DATAFILE')) as datafile:
        d = csv.Sniffer().sniff(datafile.read(1024))
        datafile.seek(0)
        dictReader = csv.DictReader(datafile, dialect=d)
        return list(dictReader)


def _get_stat_info(source):
    output = {}
    fat_avg = float(sum(get_int_from_str(
        r['Fatalities']) for r in source)) / len(source)
    aboard_sum = sum(get_int_from_str(r['Aboard']) for r in source)
    output['title'] = 'Statistical insights:'
    output[1] = ['Average fatalities for all accidents', round(fat_avg, 2)]
    output[2] = ['Total number of passengers aboard during all flights', aboard_sum]
    return output


def get_int_from_str(val):
    if not val:
        return 0
    return int(val)


def _get_aggreg_info(source):
    output = {}
    mapped_source = []
    for d in source:
        newdict = dict(
            map(lambda kv: (kv[0], extract_year(kv[0], kv[1])), d.items()))
        mapped_source.append(newdict)
    by_year = [(k, len(list(g)))
               for k, g in groupby(sorted(mapped_source, key=keyfn), keyfn)]
    output['title'] = 'Aggregate insights:'
    output[1] = ['Accidents by year', by_year]
    return output


def keyfn(x):
    return x['Date']


def extract_year(key, value):
    if key == 'Date':
        year = int(value[-4:])
        floored = int(math.floor(year / 10.0)) * 10
        value = f'{floored} - {floored + 10}'
    return value


def _get_summary_info(source):
    output = {}
    all_accid = len(source)
    all_airlines = count_distinct(source)['Operator']
    output['title'] = 'Summary insights:'
    output[1] = ['Total number of accidents', all_accid]
    output[2] = ['Airlines involved overall', all_airlines]
    return output


def count_distinct(arg_list):
    values_per_key = {}
    for d in arg_list:
        for k, v in d.items():
            values_per_key.setdefault(k, set()).add(v)
    return {k: len(v) for k, v in values_per_key.items()}


def _print_report(report):
    print(report['title'])
    for key, value in report.items():
        if key != 'title':
            print(f'{value[0]}: {value[1]}')


def _generate_xls_report(report):
    f = openpyxl.styles.Font(color="0000FF", bold=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    set_xls_styling(sheet)
    for part in report:
        entrance_check = sheet._current_row == 1
        title = sheet.cell(
            row=1 if entrance_check else sheet._current_row + 1, column=1, value=part['title'])
        title.font = f
        title.border = XLS_BORDER
        sheet.merge_cells(start_row=sheet._current_row,
                          start_column=1, end_row=sheet._current_row, end_column=2)
        process_report_part(part, sheet)
    if os.environ.get('XLS_NAME'):
        workbook.save(os.environ.get('XLS_NAME'))
    else:
        now = datetime.now()
        workbook.save(f'report-{now.strftime("%d-%m-%Y-%H_%M_%S")}.xlsx')


def process_report_part(part, sheet):
    part_item = 0
    max_row = sheet._current_row + len(part) - 1
    aggr_check = part['title'] == 'Aggregate insights:'
    if aggr_check:
        max_row += len(part[1][1])
    for row in sheet.iter_rows(min_row=sheet._current_row + 1, max_col=2, max_row=max_row):
        part_item += 1
        if not part_item in part and not aggr_check:
            break
        for cell in row:
            if aggr_check:
                if part_item > 1:
                    pair = part[1][1][part_item - 2]
                    cell.value = pair[0] if cell.column == 1 else pair[1]
                elif part_item == 1:
                    cell.value = part[1][0]
                    cell.alignment = Alignment(horizontal='center')
                    sheet.merge_cells(
                        start_row=sheet._current_row, start_column=1, end_row=sheet._current_row, end_column=2)
            else:
                cell.value = part[part_item][0] if cell.column == 1 else part[part_item][1]
            cell.border = XLS_BORDER


def set_xls_styling(sheet):
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 60
    sheet.title = "Report"
    sheet.cell(row=1, column=1)


if __name__ == '__main__':
    run()
